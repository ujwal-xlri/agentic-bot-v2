import io
from dataclasses import dataclass
from typing import List


@dataclass
class FailedFileRecord:
    filename: str
    reason: str


def build_failed_records(summary: dict) -> List[FailedFileRecord]:
    """Extract FailedFileRecord list from an ingest_folder() summary dict."""
    records = []
    for filename, info in summary.items():
        if info.get("added", 0) == 0:
            reason = info.get("error") or "No text could be extracted (image-only or empty document)"
            records.append(FailedFileRecord(filename=filename, reason=reason))
    return records


def generate_failed_files_excel(records: List[FailedFileRecord]) -> bytes:
    """Return an Excel workbook as bytes with one row per failed file."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Files"

    ws["A1"] = "File Name"
    ws["B1"] = "Reason for Not Ingesting"
    for cell in (ws["A1"], ws["B1"]):
        cell.font = Font(bold=True)

    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=record.filename)
        ws.cell(row=row_idx, column=2, value=record.reason)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
